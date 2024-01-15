import csv
import openpyxl

from datetime import datetime, time

from django.shortcuts import render, redirect
from django.views import View

from .forms import LoginForm, TankForm
from django.contrib.auth import authenticate, login
from rest_framework.decorators import api_view
from .models import CustomUser, vehicle, transactions, pumpInfo
from rest_framework.response import Response
from rest_framework import status
from .serializers import TankSerializer, TransactionSerializer, PumpSerializer
from rest_framework.views import APIView

from django.contrib import messages

# Create your views here.
from django.http import HttpResponse


def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user = authenticate(request, email=email, password=password)
            if user is not None:
                login(request, user)
                return redirect('create')
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})


@api_view(['POST'])
def getData(request):
    # Get the username and password from the request data
    email = request.data.get('email')
    password = request.data.get('password')

    # Perform user authentication here (e.g., check the credentials against your database)
    try:
        user = CustomUser.objects.get(email=email)
        if user.check_password(password):
            # Authentication succeeded
            return Response({'message': 'Login successful'}, status=status.HTTP_200_OK)
    except CustomUser.DoesNotExist:
        pass

    # Authentication failed
    return Response({'message': 'Login failed'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
def dispense(request):
    vehicleNumber = request.data.get('vehicleNumber')
    try:
        vehicleDetails = vehicle.objects.get(vehicleNumber=vehicleNumber)
        serializer = TankSerializer(vehicleDetails)
        return Response({'Quantity': vehicleDetails.quantity})
    except vehicleDetails.DoesNotExist:
        return Response({'status': 'Vehicle does not exist'}, status=status.HTTP_404_NOT_FOUND)


def enter_vehicle_data(request):
    if request.method == 'POST':
        form = TankForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehicle data has been successfully entered.')

    vehicleDetails = vehicle.objects.all() 

    form = TankForm()

    return render(request, 'create.html', {'form': form, 'vehicle': vehicleDetails})


@api_view(['POST'])
def check_vehicle(request):
    if request.method == 'POST':
        vehicleNumber = request.data.get('vehicleNumber')
        odometer = request.data.get('odometer')

        try:
            vehicleDetails = vehicle.objects.get(vehicleNumber=vehicleNumber)

            if odometer is not None:
                # Save odometer value if provided
                vehicleDetails.odometer = odometer
                vehicleDetails.save()

            return Response({'status': 'Vehicle exists'})
        except vehicleDetails.DoesNotExist:
            return Response({'status': 'Vehicle does not exist'}, status=status.HTTP_404_NOT_FOUND)


class VehicleList(APIView):
    def get(self, request, *args, **kwargs):
        vehicles = vehicle.objects.all()
        serializer = TankSerializer(vehicles, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


# class vehicleDataView(APIView):
#     def get(self, request, vehicleNumber=None, *args, **kwargs):
#         if vehicleNumber:
#             try:
#                 vehicleDetails = vehicle.objects.get(vehicleNumber=vehicleNumber)
#                 serializer = TankdetailSerializer(vehicleDetails)
#                 return Response(serializer.data, status=status.HTTP_200_OK)
#             except vehicleDetails.DoesNotExist:
#                 return Response({'status': 'Vehicle does not exist'}, status=status.HTTP_404_NOT_FOUND)
#
#         vehicleDetails = vehicle.objects.all()
#         serializer = TankdetailSerializer(vehicleDetails, many=True)
#         return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
def post_vehicle(request):
    if request.method == 'POST':
        vehicleNumber = request.data.get('vehicleNumber')
        quantity = request.data.get('quantity')

        try:
            vehicleDetails = vehicle.objects.get(vehicleNumber=vehicleNumber)
            vehicleDetails.quantity = quantity
            vehicleDetails.save()

            return Response({'status': 'Quantity updated successfully'})
        except vehicleDetails.DoesNotExist:
            return Response({'status': 'Vehicle does not exist'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
def get_vehicle(request, vehicleNumber):
    if request.method == 'GET':
        try:
            vehicleDetails = vehicle.objects.get(vehicleNumber=vehicleNumber)
            return Response({'vehicleNumber': vehicleDetails.vehicleNumber, 'quantity': vehicleDetails.quantity})
        except vehicleDetails.DoesNotExist:
            return Response({'status': 'Vehicle does not exist'}, status=status.HTTP_404_NOT_FOUND)


class transaction(View):
    template_name = 'transaction.html'

    def get(self, request):
        vehicle_details = transactions.objects.all()
        return render(request, self.template_name, {'vehicle_details': vehicle_details})

def export_csv(request):
    response = HttpResponse(content_type='text/csv')

    response['Content-Disposition'] = 'attachment; filename="transactions-'+str(datetime.now())+".csv"

    writer = csv.writer(response)
    writer.writerow(['Sr.No', 'Vehicle Number', 'Odometer', 'dispensedQuantity', 'TimeStamp'])

    for obj in transactions.objects.all():
        writer.writerow([obj.id, obj.vehicleNumber, obj.odometer, obj.dispensedQuantity, obj.timestamp])

    return response

def export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transactions-'+str(datetime.now())+".xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'My Data'

    # Write header row
    header = ['Sr.No', 'Vehicle Number', 'Odometer', 'Gallon Limit', 'TimeStamp']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = transactions.objects.all().values_list('id', 'vehicleNumber', 'odometer','gallonLimit','timestamp' )
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num+1, column=col_num)
            cell.value = str(cell_value)

    workbook.save(response)

    return response


@api_view(['POST'])
def dispensed_quantity(request):
    if request.method == 'POST':
        # Deserialize the request data using a serializer
        serializer = TransactionSerializer(data=request.data)
        if serializer.is_valid():
            # Save data to the database
            transaction = serializer.save()

            # You can return the saved data or a success message
            response_data = {

                'vehicleNumber': transaction.vehicleNumber,
                'dispensedQuantity': transaction.dispensedQuantity,
                'odometer': transaction.odometer,

            }

            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            return Response({'error': 'Invalid input'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({'error': 'Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


class getBuyDate(APIView):
    def get(self, request, date, *args, **kwargs):
        try:
            date_object = datetime.strptime(date, '%Y-%m-%d')
            transactions_for_date = transactions.objects.filter(timestamp__date=date_object.date())

            serializer = TransactionSerializer(transactions_for_date, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ValueError:
            return Response({"error": "Invalid date format. Use 'YYYY-MM-DD'."},
                            status=status.HTTP_400_BAD_REQUEST)

# class odometer(APIView):
#     def post(self, request, *args, **kwargs):
#         vehicle_number = request.data.get('vehicleNumber', None)
#         odometer_value = request.data.get('odometer', None)
#
#         if not vehicle_number or not odometer_value:
#             return Response({"error": "Both 'vehicleNumber' and 'odometer' are required."},
#                             status=status.HTTP_400_BAD_REQUEST)
#
#         try:
#             transaction = transactions.objects.get(vehicleNumber=vehicle_number)
#             transaction.odometer = odometer_value
#             transaction.save()
#
#             serializer = TransactionSerializer(transaction)
#             return Response(serializer.data, status=status.HTTP_200_OK)
#         except transactions.DoesNotExist:
#             return Response({"error": f"Transaction with vehicleNumber {vehicle_number} does not exist."},
#                             status=status.HTTP_404_NOT_FOUND)


class TransactionsBulkCreateView(APIView):
    def post(self, request, *args, **kwargs):
        data = request.data.get('transactions', [])  # Assuming your data is passed as a list under 'transactions' key

        serializer = TransactionSerializer(data=data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PumpInfo(APIView):
    def post(self, request):
        serializer = PumpSerializer(data=request.data)
        if serializer.is_valid():
            pump_number = serializer.validated_data['pumpNumber']
            pump_info, created = pumpInfo.objects.get_or_create(pumpNumber=pump_number)

            pump_info.vehicleNumber = serializer.validated_data.get('vehicleNumber', pump_info.vehicleNumber)
            pump_info.odometer = serializer.validated_data.get('odometer', pump_info.odometer)
            pump_info.pumpStatus = serializer.validated_data.get('pumpStatus', pump_info.pumpStatus)

            pump_info.save()

            return Response(PumpSerializer(pump_info).data,
                            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PumpInfoList(APIView):
    def get(self, request, format=None):
        pump_infos = pumpInfo.objects.all()
        serializer = PumpSerializer(pump_infos, many=True)
        return Response(serializer.data)